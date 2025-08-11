try:
    import fitz  # PyMuPDF
    PYMUPDF_AVAILABLE = True
except ImportError:
    PYMUPDF_AVAILABLE = False
    print("Warning: PyMuPDF not available")

try:
    import pdfplumber
    PDFPLUMBER_AVAILABLE = True
except ImportError:
    PDFPLUMBER_AVAILABLE = False
    print("Warning: pdfplumber not available")

try:
    import pandas as pd
    PANDAS_AVAILABLE = True
except ImportError:
    PANDAS_AVAILABLE = False
    print("Warning: pandas not available")
from pathlib import Path
import logging
import re
import uuid
import os
from datetime import datetime
from typing import List, Dict, Any, Optional, Tuple
import json
import asyncio
import openpyxl  # For Excel processing

from ..models.document import (
    Document, DocumentChunk, DocumentMetadata, DocumentType, 
    FinancialMetrics, ConfidenceLevel
)
from ..core.config import settings
from .ollama_service import OllamaService
from .performance_attribution_service import PerformanceAttributionService

logger = logging.getLogger(__name__)

class FinancialDocumentProcessor:
    def __init__(self, ollama_service: OllamaService):
        self.ollama_service = ollama_service
        self.performance_attribution_service = PerformanceAttributionService()
        self.chunk_size = 1000
        self.chunk_overlap = 200
        
        # Financial keywords for detection
        self.financial_keywords = {
            'revenue': ['revenue', 'sales', 'turnover', 'income', 'gross receipts'],
            'profit': ['profit', 'earnings', 'net income', 'ebitda', 'ebit'],
            'assets': ['assets', 'total assets', 'current assets', 'fixed assets'],
            'liabilities': ['liabilities', 'debt', 'obligations', 'payables'],
            'equity': ['equity', 'shareholders equity', 'stockholders equity'],
            'cash_flow': ['cash flow', 'operating cash flow', 'free cash flow'],
            'ratios': ['roe', 'roa', 'debt to equity', 'current ratio', 'quick ratio'],
            'performance': ['performance', 'attribution', 'benchmark', 'alpha', 'beta']
        }
        
        # Asset class detection keywords
        self.equity_keywords = [
            'sector', 'gics', 'industry', 'security selection', 
            'stock', 'equity', 'market cap', 'style', 'value', 'growth',
            'large cap', 'small cap', 'mid cap', 'sector allocation',
            'stock picking', 'security specific', 'share price',
            'dividend yield', 'earnings per share', 'p/e ratio'
        ]
        
        self.fixed_income_keywords = [
            'duration', 'credit', 'currency', 'country', 'sovereign',
            'corporate', 'government', 'bond', 'yield', 'fx selection',
            'interest rate', 'credit quality', 'credit spread',
            'yield curve', 'maturity', 'coupon', 'fixed income',
            'treasury', 'municipal', 'high yield', 'investment grade',
            'credit rating', 'duration risk', 'convexity'
        ]
        
    async def process_document(self, file_path: str, document_type: DocumentType, metadata: Dict[str, Any] = None) -> Document:
        """Process a financial document and extract comprehensive metadata"""
        try:
            logger.info(f"Processing document: {file_path}")
            
            # Generate document ID
            document_id = str(uuid.uuid4())
            
            # Extract basic file info
            file_info = self._get_file_info(file_path)
            
            # Process content based on file type
            logger.info(f"Processing file: {file_path}")
            
            if file_path.lower().endswith('.pdf'):
                content_data = await self._process_pdf(file_path)
            elif file_path.lower().endswith('.txt'):
                content_data = await self._process_txt(file_path)
            elif file_path.lower().endswith('.docx'):
                content_data = await self._process_txt(file_path)  # Simple text processing for now
            elif file_path.lower().endswith(('.xlsx', '.xls')):
                content_data = await self._process_excel(file_path)
            else:
                logger.error(f"DEBUG: Unsupported file type detected: {file_path}")
                logger.error(f"DEBUG: File extension: {os.path.splitext(file_path)[1]}")
                raise ValueError(f"DEBUG_ERROR: Unsupported file type: {os.path.splitext(file_path)[1]}")
            
            # Extract advanced metadata
            doc_metadata = await self._extract_metadata(
                content_data, document_type, file_info, metadata
            )
            
            # Create chunks with embeddings
            chunks = await self._create_chunks(
                document_id, content_data, doc_metadata
            )
            
            # Create document
            document = Document(
                document_id=document_id,
                metadata=doc_metadata,
                chunks=chunks,
                processing_status="completed"
            )
            
            logger.info(f"Successfully processed document {document_id} with {len(chunks)} chunks")
            return document
            
        except Exception as e:
            logger.error(f"Failed to process document {file_path}: {e}")
            raise

    def _get_file_info(self, file_path: str) -> Dict[str, Any]:
        """Extract basic file information"""
        path = Path(file_path)
        return {
            'filename': path.name,
            'file_size': path.stat().st_size,
            'file_type': path.suffix.lower(),
            'created_at': datetime.fromtimestamp(path.stat().st_ctime),
            'modified_at': datetime.fromtimestamp(path.stat().st_mtime)
        }

    async def _process_pdf(self, file_path: str) -> Dict[str, Any]:
        """Process PDF with multiple extraction strategies"""
        content_data = {
            'text_content': '',
            'tables': [],
            'pages': [],
            'images': [],
            'metadata': {},
            'structure': {}
        }
        
        # Strategy 1: PyMuPDF for text and basic structure
        doc = fitz.open(file_path)
        
        for page_num in range(len(doc)):
            page = doc[page_num]
            
            # Extract text
            text = page.get_text()
            
            # Extract images
            image_list = page.get_images()
            
            # Extract text blocks with positioning
            blocks = page.get_text("dict")
            
            page_data = {
                'page_number': page_num + 1,
                'text': text,
                'images': len(image_list),
                'blocks': blocks,
                'bbox': page.rect
            }
            
            content_data['pages'].append(page_data)
            content_data['text_content'] += f"\n\n--- Page {page_num + 1} ---\n{text}"
        
        doc.close()
        
        # Strategy 2: Camelot for table extraction (disabled for now)
        # try:
        #     tables = camelot.read_pdf(file_path, pages='all', flavor='lattice')
        #     for i, table in enumerate(tables):
        #         if table.df is not None and not table.df.empty:
        #             content_data['tables'].append({
        #                 'table_id': i,
        #                 'page': table.page,
        #                 'data': table.df.to_dict('records'),
        #                 'shape': table.df.shape,
        #                 'accuracy': table.accuracy if hasattr(table, 'accuracy') else 0.0
        #             })
        # except Exception as e:
        #     logger.warning(f"Camelot table extraction failed: {e}")
        
        # Strategy 3: PDFPlumber for additional table detection
        try:
            with pdfplumber.open(file_path) as pdf:
                for page_num, page in enumerate(pdf.pages):
                    tables_on_page = page.extract_tables()
                    for table_idx, table in enumerate(tables_on_page):
                        if table:
                            df = pd.DataFrame(table[1:], columns=table[0])
                            content_data['tables'].append({
                                'table_id': f"plumber_{page_num}_{table_idx}",
                                'page': page_num + 1,
                                'data': df.to_dict('records'),
                                'shape': df.shape,
                                'extraction_method': 'pdfplumber'
                            })
        except Exception as e:
            logger.warning(f"PDFPlumber table extraction failed: {e}")
        
        return content_data

    async def _process_txt(self, file_path: str) -> Dict[str, Any]:
        """Process TXT file"""
        content_data = {
            'text_content': '',
            'tables': [],
            'pages': [],
            'images': [],
            'metadata': {},
            'structure': {}
        }
        
        try:
            with open(file_path, 'r', encoding='utf-8') as file:
                text_content = file.read()
            
            # For text files, treat the entire content as one page
            page_data = {
                'page_number': 1,
                'text': text_content,
                'images': 0,
                'blocks': [],
                'bbox': None
            }
            
            content_data['pages'].append(page_data)
            content_data['text_content'] = text_content
            
        except UnicodeDecodeError:
            # Try with different encoding
            with open(file_path, 'r', encoding='latin-1') as file:
                text_content = file.read()
            
            page_data = {
                'page_number': 1,
                'text': text_content,
                'images': 0,
                'blocks': [],
                'bbox': None
            }
            
            content_data['pages'].append(page_data)
            content_data['text_content'] = text_content
        
        return content_data

    async def _process_excel(self, file_path: str) -> Dict[str, Any]:
        """Process Excel file (both .xlsx and .xls)"""
        content_data = {
            'text_content': '',
            'tables': [],
            'pages': [],
            'images': [],
            'metadata': {},
            'structure': {}
        }
        
        try:
            logger.info(f"Starting Excel file processing: {file_path}")
            text_parts = []
            processed_sheets = 0
            
            # Try to read with pandas first (handles both .xls and .xlsx)
            try:
                excel_file = pd.ExcelFile(file_path)
                sheet_names = excel_file.sheet_names
                logger.info(f"Found {len(sheet_names)} sheets: {sheet_names}")
                
                # Process each sheet
                for sheet_idx, sheet_name in enumerate(sheet_names):
                    try:
                        # Read the sheet
                        df = pd.read_excel(excel_file, sheet_name=sheet_name)
                        logger.info(f"Processing sheet '{sheet_name}' with shape {df.shape}")
                        
                        # Check if sheet has any content
                        if df.empty or (df.shape[0] == 0):
                            logger.warning(f"Sheet '{sheet_name}' is empty, creating minimal content")
                            page_text = f"Sheet: {sheet_name}\n(Empty sheet)"
                        else:
                            # Clean the dataframe - handle NaN values
                            df = df.fillna('')
                            
                            # Convert all columns to string to avoid JSON serialization issues
                            df_str = df.astype(str)
                            
                            # Add table to content_data
                            content_data['tables'].append({
                                'table_id': f"excel_sheet_{sheet_idx}_{sheet_name}",
                                'sheet_name': sheet_name,
                                'page': sheet_idx + 1,
                                'data': df_str.to_dict('records'),
                                'shape': list(df.shape),  # Convert tuple to list for JSON serialization
                                'extraction_method': 'pandas',
                                'columns': list(df.columns)
                            })
                            
                            # Create readable text representation
                            sheet_text = f"\n\n--- Sheet: {sheet_name} ---\n"
                            # Limit text output for very large sheets
                            if df.shape[0] > 100:
                                logger.info(f"Large sheet detected ({df.shape[0]} rows), truncating to first 100 rows for text")
                                sheet_text += df_str.head(100).to_string(index=False)
                                sheet_text += f"\n... ({df.shape[0] - 100} more rows)"
                            else:
                                sheet_text += df_str.to_string(index=False)
                            
                            text_parts.append(sheet_text)
                            page_text = f"Sheet: {sheet_name}\n" + sheet_text
                            processed_sheets += 1
                        
                        # Create page entry for this sheet
                        page_data = {
                            'page_number': sheet_idx + 1,
                            'text': page_text,
                            'sheet_name': sheet_name,
                            'images': 0,
                            'blocks': [],
                            'bbox': None
                        }
                        content_data['pages'].append(page_data)
                        
                    except Exception as sheet_error:
                        logger.error(f"Error processing sheet '{sheet_name}': {sheet_error}")
                        # Create a fallback page for this sheet
                        error_text = f"Sheet: {sheet_name}\nError processing sheet: {str(sheet_error)}"
                        page_data = {
                            'page_number': sheet_idx + 1,
                            'text': error_text,
                            'sheet_name': sheet_name,
                            'images': 0,
                            'blocks': [],
                            'bbox': None
                        }
                        content_data['pages'].append(page_data)
                        continue
                
                excel_file.close()
                logger.info(f"Excel processing completed. Processed {processed_sheets}/{len(sheet_names)} sheets successfully")
                
            except Exception as pandas_error:
                logger.error(f"Pandas Excel processing failed: {pandas_error}")
                # Fallback to openpyxl for .xlsx files
                if file_path.lower().endswith('.xlsx'):
                    logger.info("Trying openpyxl fallback for .xlsx file")
                    workbook = openpyxl.load_workbook(file_path, data_only=True)
                    sheet_names = workbook.sheetnames
                    
                    for sheet_idx, sheet_name in enumerate(sheet_names):
                        worksheet = workbook[sheet_name]
                        
                        # Simple text extraction from cells
                        sheet_text = f"\n\n--- Sheet: {sheet_name} ---\n"
                        for row in worksheet.iter_rows(values_only=True):
                            if any(cell is not None for cell in row):
                                row_text = " | ".join(str(cell) if cell is not None else "" for cell in row)
                                sheet_text += row_text + "\n"
                        
                        text_parts.append(sheet_text)
                        
                        page_data = {
                            'page_number': sheet_idx + 1,
                            'text': f"Sheet: {sheet_name}\n{sheet_text}",
                            'sheet_name': sheet_name,
                            'images': 0,
                            'blocks': [],
                            'bbox': None
                        }
                        content_data['pages'].append(page_data)
                    
                    workbook.close()
                else:
                    raise pandas_error  # Re-raise for .xls files since openpyxl can't handle them
            
            # Combine all text content
            if text_parts:
                content_data['text_content'] = "\n".join(text_parts)
            else:
                content_data['text_content'] = f"Excel file: {os.path.basename(file_path)}\nNo readable content found"
            
            # Ensure we have at least one page
            if not content_data['pages']:
                page_data = {
                    'page_number': 1,
                    'text': content_data['text_content'],
                    'images': 0,
                    'blocks': [],
                    'bbox': None
                }
                content_data['pages'].append(page_data)
            
            logger.info(f"Excel processing complete: {len(content_data['tables'])} tables, {len(content_data['pages'])} pages, {len(content_data['text_content'])} chars")
            
        except Exception as e:
            logger.error(f"Critical error processing Excel file {file_path}: {e}")
            import traceback
            logger.error(f"Full traceback: {traceback.format_exc()}")
            
            # Create minimal fallback content
            content_data['text_content'] = f"Excel file: {os.path.basename(file_path)}\nProcessing failed: {str(e)}"
            page_data = {
                'page_number': 1,
                'text': content_data['text_content'],
                'images': 0,
                'blocks': [],
                'bbox': None
            }
            content_data['pages'] = [page_data]
            content_data['tables'] = []
        
        return content_data

    async def _extract_metadata(
        self, 
        content_data: Dict[str, Any], 
        document_type: DocumentType,
        file_info: Dict[str, Any],
        additional_metadata: Dict[str, Any] = None
    ) -> DocumentMetadata:
        """Extract comprehensive metadata from document content"""
        
        # Basic metadata
        metadata = {
            'filename': file_info['filename'],
            'file_size': file_info['file_size'],
            'file_type': file_info['file_type'],
            'document_type': document_type,
            'upload_timestamp': datetime.now(),
            'total_pages': len(content_data['pages']),
            'total_chunks': 0,  # Will be updated later
            'has_tables': len(content_data['tables']) > 0,
            'has_charts': sum(page['images'] for page in content_data['pages']) > 0,
            'confidence_score': 0.8,  # Base confidence
            'key_topics': [],
            'named_entities': [],
            'tags': []
        }
        
        # Extract text for analysis
        full_text = content_data['text_content']
        
        # Extract financial metrics
        financial_metrics = self._extract_financial_metrics(full_text, content_data['tables'])
        if financial_metrics:
            metadata['financial_metrics'] = financial_metrics
            metadata['has_financial_data'] = True
        
        # Extract entities using Ollama
        try:
            entities = await self.ollama_service.extract_financial_entities(full_text[:5000])
            metadata['named_entities'] = entities.get('companies', []) + entities.get('people', [])
            metadata['key_topics'] = entities.get('metrics', []) + entities.get('products', [])
        except Exception as e:
            logger.warning(f"Entity extraction failed: {e}")
        
        # Document type specific extraction
        if document_type == DocumentType.FINANCIAL_REPORT:
            metadata.update(self._extract_financial_report_metadata(full_text))
        elif document_type == DocumentType.PERFORMANCE_ATTRIBUTION:
            metadata.update(self._extract_performance_attribution_metadata(full_text))
            
            # Detect asset class for attribution documents
            asset_class = self.detect_asset_class(full_text)
            metadata['asset_class'] = asset_class
            logger.info(f"Detected asset class for attribution document: {asset_class}")
            
            # Enhanced performance attribution processing
            try:
                attribution_data = self.performance_attribution_service.extract_attribution_data_from_tables(content_data['tables'])
                if attribution_data:
                    parsed_data = self.performance_attribution_service.parse_attribution_table(attribution_data)
                    if parsed_data:
                        metadata = self.performance_attribution_service.enhance_document_metadata(metadata, parsed_data)
                        logger.info(f"Enhanced metadata with performance attribution data: {parsed_data.get('period_name', 'Unknown period')}")
            except Exception as e:
                logger.warning(f"Failed to process performance attribution data: {e}")
                
        elif document_type == DocumentType.LEGAL_CONTRACT:
            metadata.update(self._extract_legal_contract_metadata(full_text))
        elif document_type == DocumentType.COMPLIANCE_REPORT:
            metadata.update(self._extract_compliance_metadata(full_text))
        
        # Add additional metadata if provided
        if additional_metadata:
            metadata.update(additional_metadata)
        
        # Ensure all required fields are present with defaults
        required_defaults = {
            'has_financial_data': metadata.get('has_financial_data', False),
            'custom_fields': metadata.get('custom_fields', {}),
            'tags': metadata.get('tags', [])
        }
        metadata.update(required_defaults)
        
        try:
            return DocumentMetadata(**metadata)
        except Exception as e:
            logger.error(f"Error creating DocumentMetadata: {e}")
            logger.error(f"Metadata keys: {list(metadata.keys())}")
            # Create minimal metadata that should work
            minimal_metadata = {
                'filename': metadata.get('filename', 'unknown.txt'),
                'file_size': metadata.get('file_size', 1000),
                'file_type': metadata.get('file_type', '.txt'),
                'document_type': metadata.get('document_type', 'other'),
                'upload_timestamp': metadata.get('upload_timestamp', datetime.now()),
                'total_pages': metadata.get('total_pages', 1),
                'total_chunks': metadata.get('total_chunks', 0),
                'has_financial_data': False,
                'confidence_score': 0.5,
                'tags': [],
                'custom_fields': {}
            }
            return DocumentMetadata(**minimal_metadata)

    def _extract_financial_metrics(self, text: str, tables: List[Dict]) -> Optional[FinancialMetrics]:
        """Extract financial metrics from text and tables"""
        metrics = {}
        
        # Common financial metric patterns
        patterns = {
            'revenue': r'(?:revenue|sales|turnover)[\s:]*\$?([\d,]+\.?\d*)\s*(?:million|billion|thousand|m|b|k)?',
            'ebitda': r'(?:ebitda|operating profit)[\s:]*\$?([\d,]+\.?\d*)\s*(?:million|billion|thousand|m|b|k)?',
            'net_income': r'(?:net income|profit)[\s:]*\$?([\d,]+\.?\d*)\s*(?:million|billion|thousand|m|b|k)?',
            'total_assets': r'(?:total assets)[\s:]*\$?([\d,]+\.?\d*)\s*(?:million|billion|thousand|m|b|k)?',
            'total_liabilities': r'(?:total liabilities|debt)[\s:]*\$?([\d,]+\.?\d*)\s*(?:million|billion|thousand|m|b|k)?'
        }
        
        for metric, pattern in patterns.items():
            matches = re.findall(pattern, text, re.IGNORECASE)
            if matches:
                try:
                    # Take the first match and convert to float
                    value = float(matches[0].replace(',', ''))
                    metrics[metric] = value
                except ValueError:
                    continue
        
        # Extract from tables
        for table in tables:
            if 'data' in table:
                for row in table['data']:
                    for key, value in row.items():
                        if isinstance(value, str) and any(keyword in key.lower() for keyword in ['revenue', 'profit', 'income']):
                            # Try to extract numeric value
                            numbers = re.findall(r'[\d,]+\.?\d*', str(value))
                            if numbers:
                                try:
                                    metrics[key.lower().replace(' ', '_')] = float(numbers[0].replace(',', ''))
                                except ValueError:
                                    continue
        
        return FinancialMetrics(**metrics) if metrics else None

    def _extract_financial_report_metadata(self, text: str) -> Dict[str, Any]:
        """Extract metadata specific to financial reports"""
        metadata = {}
        
        # Extract fiscal year
        year_pattern = r'(?:fiscal year|fy|year ended?)[\s:]*(20\d{2})'
        years = re.findall(year_pattern, text, re.IGNORECASE)
        if years:
            metadata['fiscal_year'] = int(years[0])
        
        # Extract quarter
        quarter_pattern = r'(?:q[1-4]|quarter\s+[1-4]|first quarter|second quarter|third quarter|fourth quarter)'
        quarters = re.findall(quarter_pattern, text, re.IGNORECASE)
        if quarters:
            metadata['fiscal_quarter'] = quarters[0].upper()
        
        # Extract company name (simple heuristic)
        company_pattern = r'([A-Z][a-z]+(?:\s+[A-Z][a-z]+)*(?:\s+(?:Inc|Corp|Ltd|LLC|Company)\.?))'
        companies = re.findall(company_pattern, text)
        if companies:
            metadata['company_name'] = companies[0]
        
        return metadata

    def _extract_performance_attribution_metadata(self, text: str) -> Dict[str, Any]:
        """Extract metadata specific to performance attribution reports"""
        metadata = {}
        
        # Extract attribution period
        period_pattern = r'(?:period|attribution period)[\s:]*([a-zA-Z]+\s+20\d{2}(?:\s*-\s*[a-zA-Z]+\s+20\d{2})?)'
        periods = re.findall(period_pattern, text, re.IGNORECASE)
        if periods:
            metadata['attribution_period'] = periods[0]
        
        # Extract benchmark
        benchmark_pattern = r'(?:benchmark|index)[\s:]*([A-Z&\s\d]+)'
        benchmarks = re.findall(benchmark_pattern, text)
        if benchmarks:
            metadata['benchmark_name'] = benchmarks[0].strip()
        
        # Extract portfolio name
        portfolio_pattern = r'(?:portfolio|fund)[\s:]*([A-Z][a-zA-Z\s]+(?:Fund|Portfolio))'
        portfolios = re.findall(portfolio_pattern, text)
        if portfolios:
            metadata['portfolio_name'] = portfolios[0].strip()
        
        return metadata

    def _extract_legal_contract_metadata(self, text: str) -> Dict[str, Any]:
        """Extract metadata specific to legal contracts"""
        metadata = {}
        
        # Extract contract parties
        party_pattern = r'(?:between|party)[\s:]*([A-Z][a-zA-Z\s,]+)(?:and|&)([A-Z][a-zA-Z\s,]+)'
        parties = re.findall(party_pattern, text)
        if parties:
            metadata['contract_parties'] = [p.strip() for p in parties[0]]
        
        # Extract effective date
        date_pattern = r'(?:effective date|dated)[\s:]*([a-zA-Z]+\s+\d{1,2},?\s+20\d{2})'
        dates = re.findall(date_pattern, text, re.IGNORECASE)
        if dates:
            metadata['effective_date'] = dates[0]
        
        return metadata

    def _extract_compliance_metadata(self, text: str) -> Dict[str, Any]:
        """Extract metadata specific to compliance reports"""
        metadata = {}
        
        # Extract regulation references
        reg_pattern = r'(?:regulation|rule|section)[\s:]*([A-Z]?[\d\w\-\.]+)'
        regulations = re.findall(reg_pattern, text, re.IGNORECASE)
        if regulations:
            metadata['regulations'] = list(set(regulations[:5]))  # Top 5 unique
        
        return metadata

    def detect_asset_class(self, content: str) -> str:
        """Detect asset class from document content"""
        try:
            content_lower = content.lower()
            
            # Count equity indicators
            equity_score = sum(1 for keyword in self.equity_keywords if keyword in content_lower)
            
            # Count fixed income indicators  
            fixed_income_score = sum(1 for keyword in self.fixed_income_keywords if keyword in content_lower)
            
            logger.info(f"Asset class detection - Equity score: {equity_score}, Fixed Income score: {fixed_income_score}")
            
            # Determine asset class
            if equity_score > fixed_income_score and equity_score > 0:
                return "equity"
            elif fixed_income_score > equity_score and fixed_income_score > 0:
                return "fixed_income"
            else:
                return "unknown"
                
        except Exception as e:
            logger.error(f"Asset class detection failed: {e}")
            return "unknown"

    async def _create_chunks(
        self, 
        document_id: str, 
        content_data: Dict[str, Any], 
        metadata: DocumentMetadata
    ) -> List[DocumentChunk]:
        """Create chunks with embeddings and metadata"""
        chunks = []
        
        # Process text content
        text_chunks = self._chunk_text(content_data['text_content'])
        
        for i, chunk_text in enumerate(text_chunks):
            # Generate embedding
            try:
                embedding = await self.ollama_service.generate_embedding(chunk_text)
            except Exception as e:
                logger.warning(f"Failed to generate embedding for chunk {i}: {e}")
                embedding = [0.0] * 768  # Fallback
            
            # Detect financial content
            contains_financial = self._contains_financial_content(chunk_text)
            financial_keywords = self._extract_financial_keywords(chunk_text)
            
            # Determine page number (rough estimate)
            page_number = self._estimate_page_number(i, len(text_chunks), metadata.total_pages)
            
            chunk = DocumentChunk(
                chunk_id=str(uuid.uuid4()),  # Use proper UUID
                document_id=document_id,
                content=chunk_text.strip(),
                chunk_index=i,
                page_number=page_number,
                chunk_type="text",
                contains_financial_data=contains_financial,
                financial_keywords=financial_keywords,
                embedding=embedding,
                embedding_model=settings.EMBEDDING_MODEL,
                processed_timestamp=datetime.now(),
                confidence_score=0.8
            )
            chunks.append(chunk)
        
        # Process tables as separate chunks
        for table_idx, table in enumerate(content_data['tables']):
            if table['data']:
                # Convert table to text representation
                table_text = self._table_to_text(table)
                
                try:
                    embedding = await self.ollama_service.generate_embedding(table_text)
                except Exception as e:
                    logger.warning(f"Failed to generate embedding for table {table_idx}: {e}")
                    embedding = [0.0] * 768
                
                table_chunk = DocumentChunk(
                    chunk_id=str(uuid.uuid4()),  # Use proper UUID
                    document_id=document_id,
                    content=table_text,
                    chunk_index=len(chunks),
                    page_number=table.get('page', 1),
                    chunk_type="table",
                    contains_financial_data=True,  # Tables in financial docs usually contain financial data
                    financial_keywords=self._extract_financial_keywords(table_text),
                    table_data=table,
                    embedding=embedding,
                    embedding_model=settings.EMBEDDING_MODEL,
                    processed_timestamp=datetime.now(),
                    confidence_score=0.9
                )
                chunks.append(table_chunk)
        
        # Update metadata with chunk count
        metadata.total_chunks = len(chunks)
        
        return chunks

    def _chunk_text(self, text: str) -> List[str]:
        """Split text into overlapping chunks"""
        # Clean text
        text = re.sub(r'\s+', ' ', text).strip()
        
        chunks = []
        start = 0
        
        while start < len(text):
            end = start + self.chunk_size
            
            # Try to break at sentence boundary
            if end < len(text):
                # Look for sentence end within the last 200 characters
                sentence_end = text.rfind('.', start + self.chunk_size - 200, end)
                if sentence_end > start:
                    end = sentence_end + 1
            
            chunk = text[start:end].strip()
            if chunk:
                chunks.append(chunk)
            
            # Move start position with overlap
            start = end - self.chunk_overlap
            if start >= len(text):
                break
        
        return chunks

    def _contains_financial_content(self, text: str) -> bool:
        """Check if chunk contains financial content"""
        text_lower = text.lower()
        
        for category, keywords in self.financial_keywords.items():
            if any(keyword in text_lower for keyword in keywords):
                return True
        
        # Check for monetary amounts
        if re.search(r'\$[\d,]+(?:\.\d{2})?', text):
            return True
        
        # Check for percentages
        if re.search(r'\d+(?:\.\d+)?%', text):
            return True
        
        return False

    def _extract_financial_keywords(self, text: str) -> List[str]:
        """Extract financial keywords from text"""
        keywords = []
        text_lower = text.lower()
        
        for category, keyword_list in self.financial_keywords.items():
            for keyword in keyword_list:
                if keyword in text_lower:
                    keywords.append(keyword)
        
        return list(set(keywords))  # Remove duplicates

    def _estimate_page_number(self, chunk_index: int, total_chunks: int, total_pages: int) -> int:
        """Estimate page number for a chunk"""
        if total_chunks == 0:
            return 1
        
        # Simple estimation based on chunk position
        estimated_page = int((chunk_index / total_chunks) * total_pages) + 1
        return min(estimated_page, total_pages)

    def _table_to_text(self, table: Dict[str, Any]) -> str:
        """Convert table data to text representation"""
        if not table.get('data'):
            return ""
        
        text_lines = [f"Table from page {table.get('page', 'unknown')}:"]
        
        for i, row in enumerate(table['data'][:10]):  # Limit to first 10 rows
            if i == 0:
                # Header row
                text_lines.append(" | ".join(str(v) for v in row.values()))
                text_lines.append("-" * 50)
            else:
                text_lines.append(" | ".join(str(v) for v in row.values()))
        
        return "\n".join(text_lines)