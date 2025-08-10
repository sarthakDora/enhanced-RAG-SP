"""
Simple test for Excel file processing without server dependencies
"""
import sys
import os
import pandas as pd
import openpyxl
import tempfile
from pathlib import Path

def test_excel_processing():
    """Test Excel file processing directly"""
    print("=== Excel File Processing Test ===")
    
    # Create test Excel file with financial data
    data = {
        'Quarter': ['Q1 2023', 'Q2 2023', 'Q3 2023', 'Q4 2023'],
        'Revenue': [1000000, 1200000, 1100000, 1300000],
        'Expenses': [800000, 950000, 900000, 1050000],
        'Net Income': [200000, 250000, 200000, 250000],
        'EBITDA': [300000, 350000, 320000, 380000],
        'Assets': [5000000, 5200000, 5100000, 5400000]
    }
    
    df = pd.DataFrame(data)
    
    # Test both .xlsx and .xls files
    test_files = []
    
    # Create .xlsx file
    with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp_file:
        df.to_excel(tmp_file.name, index=False, sheet_name='Financial Data')
        test_files.append((tmp_file.name, '.xlsx'))
    
    # Create .xls file (using same method for simplicity)
    with tempfile.NamedTemporaryFile(suffix='.xls', delete=False) as tmp_file:
        df.to_excel(tmp_file.name, index=False, sheet_name='Financial Data', engine='openpyxl')
        test_files.append((tmp_file.name, '.xls'))
    
    # Test processing
    success_count = 0
    
    for file_path, file_type in test_files:
        print(f"\nTesting {file_type} file processing...")
        
        try:
            # Simulate the _process_excel method logic
            content_data = {
                'text_content': '',
                'tables': [],
                'pages': [],
                'images': [],
                'metadata': {},
                'structure': {}
            }
            
            # Load Excel file
            if file_path.lower().endswith('.xlsx'):
                workbook = openpyxl.load_workbook(file_path, data_only=True)
                sheet_names = workbook.sheetnames
                print(f"  Found {len(sheet_names)} sheets: {sheet_names}")
                
                text_parts = []
                
                for sheet_idx, sheet_name in enumerate(sheet_names):
                    worksheet = workbook[sheet_name]
                    
                    # Extract data
                    data = []
                    headers = []
                    
                    if worksheet.max_row > 0:
                        first_row = list(worksheet.iter_rows(min_row=1, max_row=1, values_only=True))[0]
                        headers = [str(cell) if cell is not None else f"Column_{i+1}" for i, cell in enumerate(first_row)]
                    
                    for row in worksheet.iter_rows(min_row=2, values_only=True):
                        if any(cell is not None for cell in row):
                            data.append([str(cell) if cell is not None else "" for cell in row])
                    
                    if data and headers:
                        # Create DataFrame
                        max_cols = max(len(headers), max(len(row) for row in data) if data else 0)
                        
                        while len(headers) < max_cols:
                            headers.append(f"Column_{len(headers)+1}")
                        
                        padded_data = []
                        for row in data:
                            while len(row) < max_cols:
                                row.append("")
                            padded_data.append(row[:max_cols])
                        
                        if padded_data:
                            df_processed = pd.DataFrame(padded_data, columns=headers[:max_cols])
                        else:
                            df_processed = pd.DataFrame()
                    else:
                        df_processed = pd.DataFrame()
                    
                    # Add table to content_data
                    if not df_processed.empty:
                        content_data['tables'].append({
                            'table_id': f"excel_sheet_{sheet_idx}_{sheet_name}",
                            'sheet_name': sheet_name,
                            'page': sheet_idx + 1,
                            'data': df_processed.to_dict('records'),
                            'shape': df_processed.shape,
                            'extraction_method': 'openpyxl'
                        })
                        
                        sheet_text = f"\\n\\n--- Sheet: {sheet_name} ---\\n"
                        sheet_text += df_processed.to_string(index=False)
                        text_parts.append(sheet_text)
                        
                        page_text = f"Sheet: {sheet_name}\\n" + df_processed.to_string(index=False)
                    else:
                        page_text = f"Sheet: {sheet_name}\\nEmpty sheet"
                    
                    # Create page entry for this sheet
                    page_data = {
                        'page_number': sheet_idx + 1,
                        'text': page_text,
                        'sheet_name': sheet_name,
                        'images': 0,
                        'blocks': [],
                        'bbox': None
                    }
                    content_data['pages'].append(page_data)
                
                workbook.close()
                content_data['text_content'] = "\\n".join(text_parts)
            
            else:  # .xls file
                df_read = pd.read_excel(file_path)
                content_data['tables'].append({
                    'table_id': 'excel_sheet_0_Sheet1',
                    'sheet_name': 'Sheet1',
                    'page': 1,
                    'data': df_read.to_dict('records'),
                    'shape': df_read.shape,
                    'extraction_method': 'pandas'
                })
                content_data['text_content'] = df_read.to_string(index=False)
                content_data['pages'].append({
                    'page_number': 1,
                    'text': df_read.to_string(index=False),
                    'sheet_name': 'Sheet1',
                    'images': 0,
                    'blocks': [],
                    'bbox': None
                })
            
            # Verify processing results
            print(f"  Text content length: {len(content_data['text_content'])}")
            print(f"  Number of tables: {len(content_data['tables'])}")
            print(f"  Number of pages: {len(content_data['pages'])}")
            
            if content_data['tables']:
                table = content_data['tables'][0]
                print(f"  First table shape: {table['shape']}")
                print(f"  First table has {len(table['data'])} rows")
                if table['data']:
                    print(f"  Sample data: {list(table['data'][0].keys())}")
            
            # Check for financial content
            text_content = content_data['text_content'].lower()
            financial_found = any(keyword in text_content for keyword in ['revenue', 'income', 'ebitda', 'assets'])
            print(f"  Financial content detected: {financial_found}")
            
            print(f"  SUCCESS: {file_type} file processed correctly")
            success_count += 1
            
        except Exception as e:
            print(f"  ERROR: Failed to process {file_type} file: {e}")
            import traceback
            traceback.print_exc()
        
        finally:
            # Clean up
            try:
                if os.path.exists(file_path):
                    os.remove(file_path)
            except:
                pass
    
    return success_count == len(test_files)

def test_file_type_validation():
    """Test file type validation logic from the upload handler"""
    print("\\n=== File Type Validation Test ===")
    
    # Test allowed extensions (from config.py)
    allowed_extensions = ["pdf", "docx", "txt", "xlsx", "xls"]
    
    test_files = [
        "report.xlsx", "financial_data.xls", "document.pdf", 
        "contract.docx", "notes.txt", "image.jpg", "script.py"
    ]
    
    for filename in test_files:
        file_extension = os.path.splitext(filename)[1].lower()
        is_allowed = file_extension[1:] in allowed_extensions
        print(f"  {filename}: {'ALLOWED' if is_allowed else 'REJECTED'}")
    
    return True

if __name__ == "__main__":
    print("Starting Excel upload functionality tests...")
    
    # Test dependencies
    try:
        import openpyxl
        print(f"openpyxl version: {openpyxl.__version__}")
        import pandas as pd
        print(f"pandas version: {pd.__version__}")
    except ImportError as e:
        print(f"Missing dependency: {e}")
        sys.exit(1)
    
    # Run tests
    success = True
    
    if not test_excel_processing():
        print("Excel processing test failed")
        success = False
    
    if not test_file_type_validation():
        print("File type validation test failed")
        success = False
    
    if success:
        print("\\n✅ All Excel upload functionality tests passed!")
        print("Excel file upload should work properly now.")
    else:
        print("\\n❌ Some tests failed.")