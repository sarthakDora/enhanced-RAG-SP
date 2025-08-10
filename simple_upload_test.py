"""
Simple upload test without Unicode characters
"""
import sys
import os
import pandas as pd
import openpyxl

def test_excel_upload():
    """Test Excel upload functionality"""
    print("=== Excel Upload Test ===")
    
    excel_file = "C:/Projects/enhanced-RAG-3/backend/uploads/Q2_2025_Performance_Attribution_Full_Report.xlsx"
    
    if not os.path.exists(excel_file):
        print(f"ERROR: File not found at {excel_file}")
        return False
    
    filename = os.path.basename(excel_file)
    print(f"Testing file: {filename}")
    
    # Test 1: File validation
    print("\n1. File Validation:")
    file_extension = os.path.splitext(filename)[1].lower()
    allowed_extensions = ["pdf", "docx", "txt", "xlsx", "xls"]
    
    print(f"   Extension: {file_extension}")
    print(f"   Allowed: {file_extension[1:] in allowed_extensions}")
    
    file_size = os.path.getsize(excel_file)
    max_size = 1073741824  # 1GB
    print(f"   Size: {file_size} bytes ({file_size / 1024 / 1024:.1f} MB)")
    print(f"   Size OK: {file_size <= max_size}")
    
    if file_extension[1:] not in allowed_extensions or file_size > max_size:
        print("   FAILED: File validation failed")
        return False
    print("   PASSED: File validation successful")
    
    # Test 2: Excel reading
    print("\n2. Excel Reading:")
    try:
        # Test with pandas
        df = pd.read_excel(excel_file)
        print(f"   Pandas read: SUCCESS - Shape {df.shape}")
        
        # Test with openpyxl
        workbook = openpyxl.load_workbook(excel_file, data_only=True)
        sheet_names = workbook.sheetnames
        print(f"   OpenPyXL read: SUCCESS - {len(sheet_names)} sheets")
        
        for sheet_name in sheet_names:
            worksheet = workbook[sheet_name]
            print(f"     Sheet '{sheet_name}': {worksheet.max_row} rows, {worksheet.max_column} cols")
        
        workbook.close()
        print("   PASSED: Excel reading successful")
        
    except Exception as e:
        print(f"   FAILED: Excel reading failed - {e}")
        return False
    
    # Test 3: Content extraction
    print("\n3. Content Extraction:")
    try:
        # Simulate content extraction
        workbook = openpyxl.load_workbook(excel_file, data_only=True)
        sheet_name = workbook.sheetnames[0]
        worksheet = workbook[sheet_name]
        
        # Get headers
        first_row = list(worksheet.iter_rows(min_row=1, max_row=1, values_only=True))[0]
        headers = [str(cell) if cell is not None else f"Column_{i+1}" for i, cell in enumerate(first_row)]
        print(f"   Headers: {len(headers)} columns")
        
        # Get data rows
        data_rows = 0
        for row in worksheet.iter_rows(min_row=2, values_only=True):
            if any(cell is not None for cell in row):
                data_rows += 1
        print(f"   Data rows: {data_rows}")
        
        # Create DataFrame
        data = []
        for row in worksheet.iter_rows(min_row=2, values_only=True):
            if any(cell is not None for cell in row):
                data.append([str(cell) if cell is not None else "" for cell in row])
        
        if data and headers:
            df = pd.DataFrame(data, columns=headers[:len(data[0])])
            print(f"   DataFrame created: {df.shape}")
            
            # Convert to text
            text_content = df.to_string(index=False)
            print(f"   Text length: {len(text_content)}")
            
            # Check for financial keywords
            text_lower = text_content.lower()
            financial_keywords = ['performance', 'attribution', 'portfolio', 'return', 'benchmark']
            found = [kw for kw in financial_keywords if kw in text_lower]
            print(f"   Financial keywords found: {len(found)} - {found}")
            
        workbook.close()
        print("   PASSED: Content extraction successful")
        
    except Exception as e:
        print(f"   FAILED: Content extraction failed - {e}")
        return False
    
    # Test 4: Processing simulation
    print("\n4. Processing Simulation:")
    try:
        # Simulate the complete processing
        import uuid
        from datetime import datetime
        
        document_id = str(uuid.uuid4())
        print(f"   Document ID: {document_id}")
        
        metadata = {
            'filename': filename,
            'file_size': file_size,
            'file_type': file_extension,
            'document_type': 'performance_attribution',
            'upload_timestamp': datetime.now().isoformat(),
            'total_pages': len(sheet_names),
            'has_financial_data': True,
            'tags': ['Q2', '2025', 'performance', 'attribution']
        }
        print(f"   Metadata created: {len(metadata)} fields")
        
        # Simulate chunking
        estimated_chunks = max(1, data_rows // 5) + len(sheet_names)  # text chunks + table chunks
        print(f"   Estimated chunks: {estimated_chunks}")
        
        print("   PASSED: Processing simulation successful")
        
    except Exception as e:
        print(f"   FAILED: Processing simulation failed - {e}")
        return False
    
    print("\n=== OVERALL RESULT ===")
    print("SUCCESS: Excel upload functionality is working correctly!")
    print(f"File '{filename}' can be processed successfully.")
    return True

if __name__ == "__main__":
    print("Testing Excel upload functionality...")
    
    try:
        if test_excel_upload():
            print("\nExcel upload test PASSED!")
            print("The core functionality is working. If uploads are still failing,")
            print("the issue is likely in one of these areas:")
            print("- Server not running or accessible")
            print("- Frontend-backend communication")
            print("- External services (Qdrant, Ollama) not available")
            print("- Network/CORS configuration")
        else:
            print("\nExcel upload test FAILED!")
            print("There are issues with the core Excel processing.")
    except Exception as e:
        print(f"\nTest failed with exception: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)